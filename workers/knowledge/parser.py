#!/usr/bin/env python3
"""FDE Knowledge Worker: converts documents to JSONL."""

import json
import sys
from pathlib import Path


def parse_pdf(path: Path) -> list[dict]:
    """Extract text from PDF, one record per page."""
    try:
        import pdfplumber
    except ImportError:
        return [{"error": "pdfplumber not installed"}]
    records = []
    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            if text.strip():
                records.append({
                    "source_ref": f"{path.name}#page-{i+1}",
                    "content": text.strip(),
                })
    return records


def parse_docx(path: Path) -> list[dict]:
    """Extract paragraphs from Word document."""
    try:
        from docx import Document
    except ImportError:
        return [{"error": "python-docx not installed"}]
    doc = Document(path)
    paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
    if not paragraphs:
        return []
    content = "\n".join(paragraphs)
    return [{"source_ref": f"{path.name}#docx", "content": content}]


def parse_markdown(path: Path) -> list[dict]:
    """Extract text from Markdown."""
    text = path.read_text(encoding="utf-8")
    if not text.strip():
        return []
    return [{"source_ref": f"{path.name}#md", "content": text.strip()}]


def parse_csv(path: Path) -> list[dict]:
    """Convert CSV to normalized table records."""
    import csv
    records = []
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for i, row in enumerate(reader):
            records.append({
                "source_ref": f"{path.name}#row-{i+1}",
                **row,
            })
    return records


def parse_excel(path: Path) -> list[dict]:
    """Convert Excel to normalized table records."""
    try:
        import openpyxl
    except ImportError:
        return [{"error": "openpyxl not installed"}]
    wb = openpyxl.load_workbook(path, read_only=True)
    records = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [str(h).strip() if h else f"col_{j}" for j, h in enumerate(rows[0])]
        for i, row in enumerate(rows[1:]):
            record = {"source_ref": f"{path.name}#{sheet_name}-row-{i+1}"}
            for j, value in enumerate(row):
                if j < len(headers):
                    record[headers[j]] = value
            records.append(record)
    wb.close()
    return records


PARSERS = {
    ".pdf": parse_pdf,
    ".docx": parse_docx,
    ".md": parse_markdown,
    ".markdown": parse_markdown,
    ".csv": parse_csv,
    ".xlsx": parse_excel,
    ".xls": parse_excel,
}


def main():
    if len(sys.argv) < 3:
        print("Usage: parser.py <input_file> <output_jsonl>", file=sys.stderr)
        sys.exit(1)
    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])

    ext = input_path.suffix.lower()
    parser = PARSERS.get(ext)
    if parser is None:
        print(f"Unsupported file type: {ext}", file=sys.stderr)
        sys.exit(1)

    records = parser(input_path)
    with open(output_path, "w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")
    print(f"Parsed {len(records)} records from {input_path.name}")
    sys.exit(0)


if __name__ == "__main__":
    main()
